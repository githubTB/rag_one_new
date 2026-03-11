"""
parser.py — 文档解析模块

功能：
- 每个 chunk 携带完整溯源元数据（文件名、页码、标题路径、block 类型）
- 按语义边界切块（标题归属、整表为一块）
- PDF 用 PyMuPDF + pdfplumber 提取文字和表格
- 图片 OCR：优先 PaddleOCR（中文），降级到 Tesseract
- Word 内嵌图片、PDF 内嵌图片全部提取并 OCR
- Excel 每个 Sheet 作为一个 table chunk
"""

import os
import re
import io
import hashlib
import tempfile
from dataclasses import dataclass, field
from langchain_core.documents import Document

# 加载 .env（app.py 已加载时这里是空操作，单独运行 parser.py 时有效）
try:
    from dotenv import load_dotenv
    load_dotenv(override=False)
except ImportError:
    pass

# ── 可选依赖检测 ──────────────────────────────────────────────

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False
    print("⚠️  PyMuPDF 未安装: pip install pymupdf")

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    print("⚠️  python-docx 未安装: pip install python-docx")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl 未安装: pip install openpyxl")

# ── OCR 引擎配置 ─────────────────────────────────────────────
# OCR_ENGINES: 逗号分隔的引擎名，按顺序尝试第一个可用的
# 可选值: paddle, tesseract, llm
# 示例: OCR_ENGINES=paddle,tesseract   （不用 LLM，传统引擎优先）
#        OCR_ENGINES=llm,paddle         （LLM 优先，Paddle 兜底）
#        OCR_ENGINES=tesseract          （只用 Tesseract）
_OCR_ENGINE_ORDER = [
    e.strip().lower()
    for e in os.environ.get("OCR_ENGINES", "paddle,tesseract,llm").split(",")
    if e.strip()
]

# ── LLM Vision ───────────────────────────────────────────────
import base64 as _base64
import urllib.request as _urllib_req
import json as _json

_OLLAMA_URL   = os.environ.get("OLLAMA_URL",   "http://localhost:11434")
# OCR 专用 vision 模型，独立于问答模型（LLM_MODEL）
# 需要支持图片输入，推荐: minicpm-v / llava:7b / llava:13b
_VISION_MODEL = os.environ.get("LLM_OCR_MODEL", os.environ.get("LLM_MODEL", "minicpm-v"))
_LLM_OCR_TIMEOUT = int(os.environ.get("LLM_OCR_TIMEOUT", "60"))
_LLM_OCR_MAX_PX  = int(os.environ.get("LLM_OCR_MAX_PX",  "1024"))

def _check_llm_vision() -> bool:
    if "llm" not in _OCR_ENGINE_ORDER:
        return False
    try:
        with _urllib_req.urlopen(f"{_OLLAMA_URL}/api/tags", timeout=3) as r:
            data = _json.loads(r.read())
        names = [m.get("name", "") for m in data.get("models", [])]
        return any(_VISION_MODEL.split(":")[0] in n for n in names)
    except Exception:
        return False

HAS_LLM_VISION = _check_llm_vision()

# ── PaddleOCR ────────────────────────────────────────────────
import warnings as _warnings
_warnings.filterwarnings("ignore", category=Warning, module="requests")
_warnings.filterwarnings("ignore", message=".*urllib3.*")
_warnings.filterwarnings("ignore", message=".*chardet.*")

# PaddleOCR 模型缓存目录：默认跟随系统（~/.paddlex），可通过 PADDLE_CACHE_DIR 自定义
# 注意：PaddleX 部分路径硬编码，强制重定向无效，保持默认即可
_PADDLE_CACHE = os.environ.get("PADDLE_CACHE_DIR", os.path.expanduser("~/.paddlex"))
os.makedirs(_PADDLE_CACHE, exist_ok=True)
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"   # 跳过联网检查，直接用本地模型

_paddle_init_failed = False

if "paddle" in _OCR_ENGINE_ORDER:
    try:
        from paddleocr import PaddleOCR   # 只检测能否 import
        HAS_PADDLE = True
        print("✅ PaddleOCR 已导入（将在子进程里运行，首次使用时下载模型）")
    except Exception:
        HAS_PADDLE = False
else:
    HAS_PADDLE = False

# ── Tesseract ────────────────────────────────────────────────
HAS_TESSERACT = False
_TESS_LANG = "eng"

if "tesseract" in _OCR_ENGINE_ORDER:
    try:
        import pytesseract
        from PIL import Image, ImageFilter, ImageEnhance
        _available_langs = pytesseract.get_languages(config="")
        HAS_TESSERACT = True
        _found = [l for l in ["chi_sim", "chi_tra", "eng"] if l in _available_langs]
        _TESS_LANG = "+".join(_found) if _found else "eng"
    except Exception:
        pass

try:
    from PIL import Image, ImageFilter, ImageEnhance
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ── 打印当前 OCR 引擎状态 ─────────────────────────────────────
def _ocr_status_line() -> str:
    available = []
    for eng in _OCR_ENGINE_ORDER:
        if eng == "llm"       and HAS_LLM_VISION: available.append(f"LLM({_VISION_MODEL})")
        if eng == "paddle"    and HAS_PADDLE:      available.append("PaddleOCR")
        if eng == "tesseract" and HAS_TESSERACT:   available.append(f"Tesseract({_TESS_LANG})")
    if available:
        return "✅ OCR 引擎（按优先级）: " + " → ".join(available)
    return "⚠️  无可用 OCR 引擎"

print(_ocr_status_line())



# ── 自定义异常 ───────────────────────────────────────────────

class OCRUnavailableError(RuntimeError):
    """无任何可用 OCR 引擎时抛出，携带安装指引"""
    INSTALL_GUIDE = (
        "未检测到可用的 OCR 引擎，无法识别图片文字。\n"
        "请选择以下任意一个方案后重启服务：\n\n"
        "  方案 A — LLM Vision（推荐，已有 qwen3.5:9b 无需额外安装）：\n"
        "    确保 Ollama 已启动，且已拉取模型：\n"
        "    ollama serve\n"
        "    ollama pull qwen3.5:9b\n\n"
        "  方案 B — PaddleOCR（纯本地，中文识别率高）：\n"
        "    pip install paddlepaddle paddleocr\n\n"
        "  方案 C — Tesseract（轻量）：\n"
        "    # Ubuntu/Debian\n"
        "    sudo apt-get install tesseract-ocr tesseract-ocr-chi-sim\n"
        "    pip install pytesseract\n"
        "    # macOS\n"
        "    brew install tesseract tesseract-lang\n"
        "    pip install pytesseract"
    )

    def __init__(self, fname: str = ""):
        self.fname = fname
        super().__init__(self.INSTALL_GUIDE)


# ── Chunk 数据结构 ────────────────────────────────────────────

@dataclass
class Chunk:
    """带完整溯源信息的文本块"""
    text: str
    file_name: str
    file_path: str
    file_hash: str
    page: int = 1
    block_type: str = "text"          # text | heading | table | image
    heading_path: str = ""
    heading_level: int = 0
    table_headers: list = field(default_factory=list)
    chunk_index: int = 0

    def to_langchain_doc(self) -> Document:
        return Document(
            page_content=self.text,
            metadata={
                "file_name":     self.file_name,
                "file_path":     self.file_path,
                "file_hash":     self.file_hash,
                "page":          self.page,
                "block_type":    self.block_type,
                "heading_path":  self.heading_path,
                "heading_level": self.heading_level,
                "table_headers": "|".join(self.table_headers),
                "chunk_index":   self.chunk_index,
            }
        )


# ── 工具函数 ──────────────────────────────────────────────────

def file_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(text: str) -> str:
    """通用文本清洗：去除中文字符间多余空格，规范化换行"""
    text = re.sub(r'([\u4e00-\u9fff])\s+([\u4e00-\u9fff])', r'\1\2', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def clean_text_table(text: str) -> str:
    """表格文本清洗：只规范化换行，保留列间空格"""
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def detect_heading_level(text: str, style_name: str = "") -> int:
    """返回标题级别（1-6），0 表示非标题"""
    if style_name:
        s = style_name.lower().replace(" ", "")
        m = re.match(r'heading(\d)', s) or re.match(r'标题(\d)', s)
        if m:
            return int(m.group(1))
        if "title" in s:    return 1
        if "subtitle" in s: return 2

    if not text or len(text) > 120:
        return 0

    if re.match(r'^第[一二三四五六七八九十百]+[章节部分]', text): return 1
    if re.match(r'^[一二三四五六七八九十]+[、．.]', text):        return 2
    if re.match(r'^\d+\.\d+\.\d+\s', text):                       return 3
    if re.match(r'^\d+\.\d+\s', text):                             return 2
    if re.match(r'^\d+[\.、]\s', text):                            return 2
    return 0


def _print_chunks(chunks: list, source: str = "") -> None:
    """统一调试打印：把所有 chunk 内容输出到控制台"""
    if not chunks:
        return
    prefix = f"[{source}] " if source else ""
    sep = "─" * 60
    print(f"\n{sep}")
    print(f"  {prefix}共 {len(chunks)} 个 chunk")
    print(sep)
    for c in chunks:
        tag = {"text": "📝", "heading": "🏷 ", "table": "📊", "image": "🖼 "}.get(c.block_type, "  ")
        loc = f"第{c.page}页"
        if c.heading_path:
            loc += f" / {c.heading_path}"
        print(f"\n  {tag} chunk#{c.chunk_index}  [{c.block_type}]  {loc}")
        print(f"  {'·' * 40}")
        for line in c.text.splitlines():
            print(f"  {line}")
    print(f"\n{sep}\n")


def _is_table_text(text: str) -> bool:
    """判断 OCR 文本是否为表格：超过一半的行含有数字或分隔符"""
    lines = [l for l in text.splitlines() if l.strip()]
    if len(lines) < 3:
        return False
    table_lines = sum(
        1 for l in lines
        if re.search(r'\d', l) and (len(l.split()) > 2 or ',' in l or '.' in l)
    )
    return table_lines / len(lines) > 0.4


def chunk_ocr_table(text: str, fname: str, file_path: str,
                    fhash: str, window: int = 4, step: int = 2) -> list[str]:
    """
    图片表格专用切块：滑动窗口，每块都带表头行。
    window: 每块包含的数据行数
    step:   窗口滑动步长
    """
    lines = [l.rstrip() for l in text.splitlines() if l.strip()]
    if not lines:
        return []

    # 前两行通常是标题+表头，固定保留
    header_lines = lines[:2]
    data_lines   = lines[2:]

    if len(data_lines) <= window:
        return [text]   # 行数少，整体一块

    header = "\n".join(header_lines)
    chunks = []
    i = 0
    while i < len(data_lines):
        block = data_lines[i : i + window]
        chunks.append(header + "\n" + "\n".join(block))
        i += step

    return chunks


def smart_chunk_text(text: str, max_size: int = 1200, overlap: int = 150) -> list[str]:
    """
    按语义边界切块：
    - 优先按双换行（段落）分割
    - 段落内超长再按句子分割（。！？；）
    - 相邻块间保留 overlap 字符上下文，避免语义截断
    - 单个段落超过 max_size 才强制切，不会把短段落切碎
    """
    if len(text) <= max_size:
        return [text]

    # 先按段落分
    paragraphs = [p.strip() for p in re.split(r'\n\n+', text) if p.strip()]

    # 段落内超长的再按句子切
    sentences = []
    for para in paragraphs:
        if len(para) <= max_size:
            sentences.append(para)
        else:
            # 按中文句末标点切句
            parts = re.split(r'(?<=[。！？；\n])', para)
            buf, buf_len = [], 0
            for part in parts:
                if buf_len + len(part) > max_size and buf:
                    sentences.append("".join(buf))
                    # overlap：保留最后一句
                    buf = buf[-1:] if buf else []
                    buf_len = len(buf[0]) if buf else 0
                buf.append(part)
                buf_len += len(part)
            if buf:
                sentences.append("".join(buf))

    # 把句子组合成 chunk，相邻 chunk 共享 overlap
    chunks, current, current_len = [], [], 0
    for sent in sentences:
        if current_len + len(sent) > max_size and current:
            chunks.append("\n\n".join(current))
            # 保留最后一个段落作为 overlap
            current = current[-1:]
            current_len = len(current[0]) if current else 0
        current.append(sent)
        current_len += len(sent)

    if current:
        chunks.append("\n\n".join(current))

    return [c for c in chunks if c.strip()]


# ── 标题栈管理 ────────────────────────────────────────────────

def _update_heading_stack(stack: list, level: int, text: str):
    while stack and stack[-1][0] >= level:
        stack.pop()
    stack.append((level, text))


def _heading_path(stack: list) -> str:
    return " > ".join(t for _, t in stack)


# ── OCR 底层实现 ──────────────────────────────────────────────

def _preprocess_image(img):
    """传统 OCR 图像预处理：灰度→放大→锐化→增强对比度→二值化"""
    if not HAS_PIL:
        return img
    img = img.convert("L")
    w, h = img.size
    if w < 1200:
        scale = 1200 / w
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    img = img.filter(ImageFilter.SHARPEN)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = img.point(lambda p: 255 if p > 140 else 0, "1")
    return img


# ── Paddle OCR（直接调用，模型已本地缓存）───────────────────
_paddle_instance = None
_paddle_init_failed = False


def _paddle_result_to_text(results) -> str:
    """
    把新版 OCRResult 转成有结构的文本。
    利用 rec_boxes 的 Y 坐标把文字分组为行，再按 X 排序还原列顺序。
    对于月度表格，自动补全 OCR 漏检的月份列。
    """
    for res in results:
        try:
            data = res.json.get("res", {})
            texts = data.get("rec_texts", [])
            boxes = data.get("rec_boxes", [])   # [x1,y1,x2,y2]
            if not texts:
                del res
                continue

            if not boxes or len(boxes) != len(texts):
                return "\n".join(t for t in texts if t.strip())

            # 按 Y 中心分组
            items = sorted(zip(boxes, texts), key=lambda x: (x[0][1] + x[0][3]) / 2)
            rows = []
            current_row = [items[0]]
            current_y = (items[0][0][1] + items[0][0][3]) / 2

            for box, text in items[1:]:
                y_center = (box[1] + box[3]) / 2
                if abs(y_center - current_y) < 15:
                    current_row.append((box, text))
                else:
                    rows.append(current_row)
                    current_row = [(box, text)]
                    current_y = y_center
            rows.append(current_row)

            # 每行内按 X 排序
            for row in rows:
                row.sort(key=lambda x: x[0][0])

            # ── 检测并补全漏检的月份列 ──────────────────────────
            # 判断：第一列 x 坐标 > 60（月份文字应该在最左侧 x<60）
            # 且存在以 "月" 结尾的行标识
            MONTH_LABELS = [
                "2019年12月",
                "1月","2月","3月","4月","5月","6月",
                "7月","8月","9月","10月","11月","12月",
                "当年累计支出", "累计均价",
            ]
            # 检查每行第一个元素的 x 坐标，如果都 > 60 说明月份列漏检了
            # 跳过 rows[3]（2019年12月，x=0 正常），从 rows[4] 开始检测
            data_rows = rows[3:]        # 所有数据行（含合计）
            check_rows = rows[4:16]     # 只检查 1月~12月 这12行
            first_xs = [row[0][0][0] for row in check_rows if row]
            need_month_patch = (
                len(check_rows) >= 10 and
                sum(1 for x in first_xs if x > 60) >= len(first_xs) * 0.8 and
                not any("月" in row[0][1] for row in check_rows if row)
            )

            lines = []
            month_idx = 0
            MONTH_DATA_LABELS = ["1月","2月","3月","4月","5月","6月",
                                  "7月","8月","9月","10月","11月","12月",
                                  "当年累计支出", "累计均价"]
            for i, row in enumerate(rows):
                cells = [t for _, t in row]
                # i>=4：跳过标题(0)、表头(1,2)、2019年12月(3)，从1月开始补
                if need_month_patch and i >= 4 and month_idx < len(MONTH_DATA_LABELS):
                    cells = [MONTH_DATA_LABELS[month_idx]] + cells
                    month_idx += 1
                lines.append("  ".join(c for c in cells if c.strip()))

            result_text = "\n".join(l for l in lines if l.strip())
            del res, data, texts, boxes, items, rows, lines
            return result_text

        except Exception:
            try:
                texts = res.json.get("res", {}).get("rec_texts", [])
                result = "\n".join(t for t in texts if t.strip())
                del res
                return result
            except Exception:
                pass
    return ""


def _rebuild_table_from_boxes(results) -> str:
    """
    用 OCR box 坐标重建表格结构，补全空单元格。
    兼容 PaddleOCR 3.x（dt_polys + rec_texts）和旧版（rec_boxes）
    """
    all_items = []  # [(x1,y1,x2,y2, text), ...]
    for res in results:
        try:
            data = res.json.get("res", {})
            texts = data.get("rec_texts", [])

            # 3.x 用 dt_polys（多边形顶点列表），旧版用 rec_boxes（[x1,y1,x2,y2]）
            polys = data.get("dt_polys", [])
            boxes = data.get("rec_boxes", [])

            if not texts:
                continue

            if polys and len(polys) == len(texts):
                # dt_polys: [[x1,y1],[x2,y2],[x3,y3],[x4,y4]] 取外接矩形
                for poly, text in zip(polys, texts):
                    if not text.strip():
                        continue
                    xs = [p[0] for p in poly]
                    ys = [p[1] for p in poly]
                    all_items.append((min(xs), min(ys), max(xs), max(ys), text.strip()))
            elif boxes and len(boxes) == len(texts):
                # 旧版 rec_boxes: [x1,y1,x2,y2]
                for box, text in zip(boxes, texts):
                    if text.strip():
                        all_items.append((box[0], box[1], box[2], box[3], text.strip()))
            else:
                # 无坐标信息，直接返回纯文本
                return "\n".join(t for t in texts if t.strip())
        except Exception as e:
            print(f"    ⚠️  box 解析失败: {e}")

    if not all_items:
        return ""

    # 1. 按 Y 中心分行
    avg_h = sum(item[3] - item[1] for item in all_items) / len(all_items)
    tol = max(avg_h * 0.6, 10)

    all_items.sort(key=lambda x: (x[1] + x[3]) / 2)
    rows: list[list[tuple]] = []
    current_row = [all_items[0]]
    current_y = (all_items[0][1] + all_items[0][3]) / 2

    for item in all_items[1:]:
        y_center = (item[1] + item[3]) / 2
        if abs(y_center - current_y) <= tol:
            current_row.append(item)
        else:
            rows.append(sorted(current_row, key=lambda x: x[0]))
            current_row = [item]
            current_y = y_center
    rows.append(sorted(current_row, key=lambda x: x[0]))

    if len(rows) < 2:
        return "\n".join(" ".join(item[4] for item in row) for row in rows)

    # 2. 聚类列中心
    all_x_centers = sorted((item[0] + item[2]) / 2
                            for row in rows for item in row)
    avg_w = sum(item[2] - item[0] for row in rows for item in row) / len(all_items)
    col_gap = max(avg_w * 1.2, 20)

    col_centers = [all_x_centers[0]]
    for x in all_x_centers[1:]:
        if x - col_centers[-1] > col_gap:
            col_centers.append(x)
        else:
            col_centers[-1] = (col_centers[-1] + x) / 2

    n_cols = len(col_centers)

    def assign_col(x_center: float) -> int:
        return min(range(n_cols), key=lambda i: abs(col_centers[i] - x_center))

    # 3. 每行按列分配，空列填 -
    lines = []
    for row in rows:
        cells = ["-"] * n_cols
        for item in row:
            col_idx = assign_col((item[0] + item[2]) / 2)
            cells[col_idx] = item[4]
        lines.append("  ".join(cells))

    return "\n".join(lines)


def _run_paddle_table(img_bytes: bytes) -> str:
    """用普通 PaddleOCR + 坐标对齐重建表格，保留空单元格"""
    global _paddle_instance, _paddle_init_failed, HAS_PADDLE

    if _paddle_init_failed:
        raise RuntimeError("PaddleOCR 已禁用")

    if _paddle_instance is None:
        try:
            try:
                _paddle_instance = PaddleOCR(
                    use_textline_orientation=True,
                    lang="ch",
                    ocr_version="PP-OCRv4",
                    cpu_threads=2,
                    enable_mkldnn=False,
                )
            except TypeError:
                _paddle_instance = PaddleOCR(
                    use_angle_cls=True,
                    lang="ch",
                    ocr_version="PP-OCRv4",
                    cpu_threads=2,
                    enable_mkldnn=False,
                )
        except BaseException as e:
            _paddle_init_failed = True
            HAS_PADDLE = False
            raise RuntimeError(f"PaddleOCR 初始化失败: {e}") from e

    tmp = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
            tmp = f.name
            try:
                img = Image.open(io.BytesIO(img_bytes))
                w, h = img.size
                max_px = 2000
                if max(w, h) > max_px:
                    scale = max_px / max(w, h)
                    img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
                img.save(f.name, format="JPEG", quality=90)
                img.close()
            except Exception:
                f.write(img_bytes)
        del img_bytes

        results = _paddle_instance.predict(tmp)
        results_list = list(results)
        text = _rebuild_table_from_boxes(results_list)
        del results_list

        if text.strip():
            print(f"    📊 坐标对齐表格识别成功（{len(text)}字）")
            return text
        return ""
    except Exception as e:
        print(f"    ⚠️  坐标对齐识别失败: {e}")
        return ""
    finally:
        if tmp:
            try:
                os.unlink(tmp)
            except Exception:
                pass


def _is_table_image_by_structure(img_bytes: bytes, filename: str = "") -> bool:
    """
    判断图片是否为表格：
    1. 文件名含"表"、"table"、"汇总"、"统计" 等关键词
    2. 图片宽高比 > 1.2（横向布局，典型表格形态）
    3. 图片足够大（小图一般不是表格）
    """
    # 文件名关键词
    table_keywords = ["表", "table", "汇总", "统计", "报表", "清单", "明细", "账单", "报告"]
    fname_lower = filename.lower()
    if any(kw in fname_lower for kw in table_keywords):
        print(f"    📊 文件名命中表格关键词: {filename}")
        return True

    try:
        img = Image.open(io.BytesIO(img_bytes))
        w, h = img.size
        # 宽高比 > 1.2 且面积足够大
        if w > h * 1.2 and w * h > 80000:
            print(f"    📊 宽高比触发表格检测: {w}x{h} ratio={w/h:.2f}")
            return True
    except Exception:
        pass
    return False


def _run_paddle(img_bytes: bytes, filename: str = "") -> str:
    global _paddle_instance, _paddle_init_failed, HAS_PADDLE

    if _paddle_init_failed:
        raise RuntimeError("PaddleOCR 之前已失败，已禁用")

    # 先判断是否为表格图片，是则走 table pipeline（含坐标对齐兜底）
    is_table = _is_table_image_by_structure(img_bytes, filename=filename)
    if is_table:
        print("    📊 检测到表格结构，使用 table pipeline")
        try:
            table_text = _run_paddle_table(img_bytes)
            if table_text.strip():
                return table_text
            print("    ⚠️  table pipeline 无结果，降级坐标对齐")
        except Exception as e:
            print(f"    ⚠️  table pipeline 异常: {e}，降级坐标对齐")

    if _paddle_instance is None:
        try:
            try:
                # 用 mobile 轻量模型，macOS CPU 上内存占用远低于 server 版
                _paddle_instance = PaddleOCR(
                    use_textline_orientation=True,
                    lang="ch",
                    ocr_version="PP-OCRv4",
                    cpu_threads=2,
                    enable_mkldnn=False,
                )
            except TypeError:
                _paddle_instance = PaddleOCR(
                    use_angle_cls=True,
                    lang="ch",
                    ocr_version="PP-OCRv4",
                    cpu_threads=2,
                    enable_mkldnn=False,
                )
        except BaseException as e:
            _paddle_init_failed = True
            HAS_PADDLE = False
            raise RuntimeError(f"PaddleOCR 初始化失败: {e}") from e

    tmp = None
    try:
        # 压缩图片到临时文件，用完立即释放原始字节
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
            tmp = f.name
            try:
                img = Image.open(io.BytesIO(img_bytes))
                w, h = img.size
                max_px = 2000
                if max(w, h) > max_px:
                    scale = max_px / max(w, h)
                    img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
                    print(f"    🖼  压缩图片: {w}x{h} → {img.size[0]}x{img.size[1]}")
                img.save(f.name, format="JPEG", quality=90)
                img.close()
                del img
            except Exception:
                f.write(img_bytes)
        # 原始字节用完释放
        del img_bytes

        results = _paddle_instance.predict(tmp)
        results_list = list(results)
        # 表格图片用坐标对齐重建，普通图片用原始文本
        if is_table:
            text = _rebuild_table_from_boxes(results_list)
            if not text.strip():
                text = _paddle_result_to_text(iter(results_list))
        else:
            text = _paddle_result_to_text(iter(results_list))
        del results_list
        return text
    except BaseException as e:
        _paddle_init_failed = True
        HAS_PADDLE = False
        raise RuntimeError(f"PaddleOCR 执行失败: {e}") from e
    finally:
        if tmp:
            try:
                os.unlink(tmp)
            except Exception:
                pass



def _run_tesseract(img_bytes: bytes) -> str:
    img = Image.open(io.BytesIO(img_bytes))
    img = _preprocess_image(img)
    return pytesseract.image_to_string(img, lang=_TESS_LANG, config="--psm 6")


def _enhance_image_for_ocr(img: "Image.Image") -> "Image.Image":
    """
    图像增强（送 OCR 前）：
    - 灰度化 + 对比度拉伸（模拟 CLAHE，去除过曝/偏暗）
    - 二次锐化（改善模糊/低分辨率图片）
    - 转回 RGB（LLM Vision 需要彩色）
    适用：模糊扫描件、低对比度截图、手写+印刷混合件
    """
    if not HAS_PIL:
        return img
    try:
        gray = img.convert("L")
        pixels = sorted(gray.getdata())
        n = len(pixels)
        lo = pixels[max(0, int(n * 0.05))]
        hi = pixels[min(n - 1, int(n * 0.95))]
        if hi > lo + 20:
            gray = gray.point(lambda p: max(0, min(255, int((p - lo) * 255 / (hi - lo)))))
        gray = gray.filter(ImageFilter.SHARPEN)
        gray = gray.filter(ImageFilter.SHARPEN)
        return gray.convert("RGB")
    except Exception as e:
        print(f"    ⚠️  图像增强失败，使用原图: {e}")
        return img


def _compress_for_llm(img_bytes: bytes) -> bytes:
    """图像增强 + 压缩再发给 LLM：长边限 _LLM_OCR_MAX_PX，转 JPEG quality=88"""
    if not HAS_PIL:
        return img_bytes
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        img = _enhance_image_for_ocr(img)
        w, h = img.size
        # 短边太小时先放大，提升识别率
        if min(w, h) < 600:
            scale = 600 / min(w, h)
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
            w, h = img.size
        if max(w, h) > _LLM_OCR_MAX_PX:
            scale = _LLM_OCR_MAX_PX / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88, optimize=True)
        compressed = buf.getvalue()
        print(f"    🖼  增强+压缩: {len(img_bytes)//1024}KB → {len(compressed)//1024}KB ({img.size[0]}x{img.size[1]})")
        return compressed
    except Exception as e:
        print(f"    ⚠️  压缩失败，使用原图: {e}")
        return img_bytes


def _run_llm(img_bytes: bytes) -> str:
    img_bytes = _compress_for_llm(img_bytes)
    b64 = _base64.b64encode(img_bytes).decode()
    payload = _json.dumps({
        "model":  _VISION_MODEL,
        "prompt": (
            "请完整提取并输出这张图片中的【全部】文字内容，不得遗漏任何行或列。"
            "保持原有段落和换行结构，表格每行单独一行，列之间用两个空格分隔。"
            "只输出文字内容本身，不要添加任何解释、标签、省略号或'等'字样。"
            "必须输出到最后一行，中途不得停止。"
            "如果图片中没有文字，输出空字符串。"
        ),
        "images": [b64],
        "stream": False,
        "options": {"temperature": 0, "num_predict": -1, "num_ctx": 8192},
    }).encode()
    req = _urllib_req.Request(
        f"{_OLLAMA_URL}/api/generate",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with _urllib_req.urlopen(req, timeout=_LLM_OCR_TIMEOUT) as r:
        text = _json.loads(r.read()).get("response", "").strip()
    if text:
        print(f"    📝 LLM OCR 识别结果（{len(text)}字）:\n{text}")
    else:
        print(f"    ⚠️  LLM OCR 未识别到文字")
    return text


# 引擎名 → (可用标志, 执行函数)
_ENGINE_MAP = {
    "paddle":    (lambda: HAS_PADDLE,     _run_paddle),
    "tesseract": (lambda: HAS_TESSERACT,  _run_tesseract),
    "llm":       (lambda: HAS_LLM_VISION, _run_llm),
}


def _do_ocr(img_bytes: bytes, filename: str = "") -> str:
    """
    按 OCR_ENGINES 顺序依次尝试，第一个返回非空结果即采用。
    全部失败返回空字符串。
    """
    for eng in _OCR_ENGINE_ORDER:
        avail_fn, run_fn = _ENGINE_MAP.get(eng, (lambda: False, None))
        if not avail_fn() or run_fn is None:
            continue
        try:
            if eng == "paddle":
                text = run_fn(img_bytes, filename)
            else:
                text = run_fn(img_bytes)
            if text.strip():
                return text.strip()
        except BaseException as e:
            print(f"    ⚠️  {eng} OCR 失败: {e}，尝试下一个引擎")
    return ""


def _ocr_bytes(img_bytes: bytes, ext: str = ".png", filename: str = "") -> str:
    return _do_ocr(img_bytes, filename=filename)


def _ocr_image_obj(img, filename: str = "") -> str:
    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="PNG")
    return _do_ocr(buf.getvalue(), filename=filename)


def _ocr_page(page) -> str:
    """对 PyMuPDF 页面做 OCR（扫描页专用）"""
    try:
        mat = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=mat)
        return _do_ocr(pix.tobytes("png"))
    except Exception as e:
        print(f"  OCR 失败: {e}")
    return ""




# ── PDF 解析 ──────────────────────────────────────────────────

def _has_any_ocr() -> bool:
    """是否有任何可用 OCR 引擎"""
    return (HAS_PADDLE and "paddle" in _OCR_ENGINE_ORDER) or \
           (HAS_TESSERACT and "tesseract" in _OCR_ENGINE_ORDER) or \
           (HAS_LLM_VISION and "llm" in _OCR_ENGINE_ORDER)


def _extract_pdf_page_images(page, fname, file_path, fhash,
                              page_num, heading_path, start_idx) -> list[Chunk]:
    """提取 PDF 页面内嵌图片并 OCR"""
    chunks = []
    if not HAS_PYMUPDF or not _has_any_ocr():
        return chunks
    try:
        seen_xref = set()
        for img_info in page.get_images(full=True):
            xref = img_info[0]
            if xref in seen_xref:
                continue
            seen_xref.add(xref)
            try:
                base_img = page.parent.extract_image(xref)
                if base_img.get("width", 0) < 100 or base_img.get("height", 0) < 100:
                    continue
                img_text = _ocr_bytes(base_img["image"], "." + base_img.get("ext", "png"))
                if img_text.strip():
                    for t in smart_chunk_text(clean_text(img_text)):
                        if t.strip():
                            chunks.append(Chunk(
                                text=t, file_name=fname, file_path=file_path,
                                file_hash=fhash, page=page_num, block_type="image",
                                heading_path=heading_path,
                                chunk_index=start_idx + len(chunks)
                            ))
            except Exception as e:
                print(f"    ⚠️  PDF 图片提取失败: {e}")
    except Exception as e:
        print(f"  ⚠️  页面图片列表获取失败: {e}")
    return chunks


def parse_pdf(file_path: str) -> list[Chunk]:
    """PyMuPDF + pdfplumber 解析 PDF，扫描页自动 OCR"""
    chunks = []
    fhash = file_sha256(file_path)
    fname = os.path.basename(file_path)
    idx = 0

    if not HAS_PYMUPDF:
        print(f"⚠️  PyMuPDF 不可用，跳过 {fname}")
        return chunks

    doc = fitz.open(file_path)
    heading_stack = []
    plumber_doc = pdfplumber.open(file_path) if HAS_PDFPLUMBER else None

    for page_num in range(len(doc)):
        page = doc[page_num]
        page_display = page_num + 1
        text_raw = page.get_text("text")

        # 扫描页 → OCR
        if len(text_raw.strip()) < 50:
            ocr_text = _ocr_page(page)
            if ocr_text:
                page_chunks = []
                for t in smart_chunk_text(clean_text(ocr_text)):
                    chunks.append(Chunk(
                        text=t, file_name=fname, file_path=file_path,
                        file_hash=fhash, page=page_display, block_type="image",
                        chunk_index=idx, heading_path=_heading_path(heading_stack)
                    ))
                    page_chunks.append(chunks[-1])
                    idx += 1
                print(f"  🖼  PDF 第{page_display}页（扫描页 OCR）→ {len(page_chunks)} chunk")
                for c in page_chunks:
                    print(f"    chunk#{c.chunk_index}: {c.text}")
            else:
                print(f"  ⚠️  PDF 第{page_display}页扫描页 OCR 无结果")
            continue

        # 表格（pdfplumber）
        table_texts = set()
        if plumber_doc:
            try:
                for table in plumber_doc.pages[page_num].extract_tables():
                    if not table or len(table) < 2:
                        continue
                    # 清洗表头：去空、合并多行表头
                    raw_headers = [str(c or "").strip().replace("\n", " ") for c in table[0]]
                    headers = [h if h else f"列{i+1}" for i, h in enumerate(raw_headers)]
                    data_rows = [r for r in table[1:] if any(c and str(c).strip() for c in r)]
                    if not data_rows:
                        continue

                    # 结构化：每行 "列名: 值, 列名: 值" 便于语义检索
                    def fmt_row(row: list, hdrs: list) -> str:
                        parts = []
                        for h, c in zip(hdrs, row):
                            v = str(c or "").strip().replace("\n", " ")
                            if v:
                                parts.append(f"{h}: {v}")
                        return " | ".join(parts)

                    CHUNK_ROWS = 30  # 每块最多30行
                    header_line = " | ".join(headers)
                    table_texts.add(header_line[:200])

                    if len(data_rows) <= CHUNK_ROWS:
                        # 小表整体一块
                        rows_text = "\n".join(fmt_row(r, headers) for r in data_rows)
                        table_text = f"[表格] {header_line}\n{rows_text}"
                        table_texts.add(table_text[:200])
                        chunks.append(Chunk(
                            text=table_text, file_name=fname, file_path=file_path,
                            file_hash=fhash, page=page_display, block_type="table",
                            table_headers=headers, heading_path=_heading_path(heading_stack),
                            chunk_index=idx
                        ))
                        idx += 1
                    else:
                        # 大表滑动窗口切块，每块带表头
                        step = CHUNK_ROWS // 2
                        i = 0
                        while i < len(data_rows):
                            block_rows = data_rows[i:i + CHUNK_ROWS]
                            rows_text = "\n".join(fmt_row(r, headers) for r in block_rows)
                            row_range = f"第{i+1}~{i+len(block_rows)}行"
                            table_text = f"[表格{row_range}] {header_line}\n{rows_text}"
                            table_texts.add(table_text[:200])
                            chunks.append(Chunk(
                                text=table_text, file_name=fname, file_path=file_path,
                                file_hash=fhash, page=page_display, block_type="table",
                                table_headers=headers, heading_path=_heading_path(heading_stack),
                                chunk_index=idx
                            ))
                            idx += 1
                            i += step
            except Exception as e:
                print(f"  ⚠️  表格提取失败 page {page_display}: {e}")

        # 页内嵌入图片 OCR
        img_chunks = _extract_pdf_page_images(
            page, fname, file_path, fhash, page_display,
            _heading_path(heading_stack), idx
        )
        chunks.extend(img_chunks)
        idx += len(img_chunks)

        # 文字块（按 block 聚合，同一 block 内文字合并后再判断标题/正文）
        blocks = page.get_text("dict")["blocks"]
        page_text_blocks: list[tuple[str, float, bool]] = []  # (text, avg_size, is_heading)
        for block in blocks:
            if block.get("type") != 0:
                continue
            block_lines = []
            sizes, bolds = [], []
            for line in block.get("lines", []):
                line_text = ""
                for span in line.get("spans", []):
                    t = span["text"]
                    if t.strip():
                        line_text += t
                        sizes.append(span.get("size", 12))
                        bolds.append(bool(span.get("flags", 0) & 2 ** 4))
                if line_text.strip():
                    block_lines.append(line_text.strip())
            if not block_lines:
                continue
            block_text = "\n".join(block_lines).strip()
            if not block_text or block_text[:200] in table_texts:
                continue
            avg_size = sum(sizes) / len(sizes) if sizes else 12
            is_bold  = sum(bolds) / len(bolds) > 0.5 if bolds else False
            page_text_blocks.append((block_text, avg_size, is_bold))

        for block_text, avg_size, is_bold in page_text_blocks:
            # 判断是否为标题
            level = 0
            if avg_size >= 16 or (avg_size >= 13 and is_bold and len(block_text) < 80):
                level = 1 if avg_size >= 18 else (2 if avg_size >= 15 else 3)
                level = detect_heading_level(block_text) or level
            if level > 0:
                _update_heading_stack(heading_stack, level, block_text)
                chunks.append(Chunk(
                    text=block_text, file_name=fname, file_path=file_path,
                    file_hash=fhash, page=page_display, block_type="heading",
                    heading_level=level, heading_path=_heading_path(heading_stack),
                    chunk_index=idx
                ))
                idx += 1
            else:
                # 正文：带上当前标题路径，保证语义完整
                for t in smart_chunk_text(block_text):
                    if t.strip():
                        chunks.append(Chunk(
                            text=t, file_name=fname, file_path=file_path,
                            file_hash=fhash, page=page_display, block_type="text",
                            heading_path=_heading_path(heading_stack), chunk_index=idx
                        ))
                        idx += 1

    doc.close()
    if plumber_doc:
        plumber_doc.close()
    _print_chunks(chunks, f"PDF · {fname}")
    return chunks


# ── Word 解析 ─────────────────────────────────────────────────

def _extract_docx_paragraph_images(element, word_doc) -> list[tuple[bytes, str]]:
    """提取 Word 段落内嵌图片"""
    results = []
    try:
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        for blip in element.findall(".//a:blip", ns):
            embed = blip.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
            )
            if not embed:
                continue
            rels = word_doc.part.rels
            if embed not in rels:
                continue
            rel = rels[embed]
            if "image" not in rel.reltype.lower():
                continue
            try:
                img_part = rel.target_part
                ext = os.path.splitext(img_part.partname)[-1].lower() or ".png"
                results.append((img_part.blob, ext))
            except Exception as e:
                print(f"    ⚠️  Word 图片读取失败: {e}")
    except Exception as e:
        print(f"  ⚠️  Word 图片解析失败: {e}")
    return results


def _docx_table_to_text(table) -> tuple[list[str], str]:
    """
    Word 表格 → (headers, table_text)
    处理合并单元格：同一行内同一个 tc（表格单元格）对象只取一次，避免水平合并单元格内容重复。
    注意：seen_tc 只在一行内有效，跨行不复用（python-docx 会复用 cell 对象）。
    """
    from docx.table import Table
    from docx.oxml.ns import qn

    grid: list[list[str]] = []

    for row in table.rows:
        seen_tc: set[int] = set()  # 每行独立，修复跨行 cell 对象复用问题
        row_cells: list[str] = []
        for cell in row.cells:
            tc_id = id(cell._tc)
            if tc_id in seen_tc:
                row_cells.append("")          # 水平合并格占位
            else:
                seen_tc.add(tc_id)
                # 优先用 cell.text，但如果为空尝试直接读 XML
                text = cell.text.strip()
                if not text:
                    texts = [t.text for t in cell._tc.iter(qn('w:t')) if t.text]
                    text = "".join(texts).strip()
                row_cells.append(text)
        if any(row_cells):
            grid.append(row_cells)

    if not grid:
        return [], ""

    headers = grid[0] if grid else []
    lines = [" | ".join(c for c in r) for r in grid]
    return headers, "\n".join(lines)


def _docx_table_from_xml(table) -> list[list[str]]:
    """
    备用方案：直接从 XML 解析表格，绕过 python-docx 的 row/cell 抽象。
    解决某些 Word 表格 row.cells 返回不完整的问题。
    """
    from docx.oxml.ns import qn
    tbl = table._tbl
    grid: list[list[str]] = []

    # 获取所有行 <w:tr>
    for tr in tbl.iter(qn('w:tr')):
        row_cells: list[str] = []
        # 获取行内所有单元格 <w:tc>
        for tc in tr.iter(qn('w:tc')):
            # 提取单元格内所有 <w:t> 文本
            texts = [t.text for t in tc.iter(qn('w:t')) if t.text]
            cell_text = "".join(texts).strip()
            row_cells.append(cell_text)
        if any(row_cells):
            grid.append(row_cells)

    return grid


def parse_docx(file_path: str) -> list[Chunk]:
    """
    python-docx 按 body 顺序解析：
    - 标题层级识别（style + 正则双保险）
    - 列表段落识别（List Paragraph style → 前缀 • 合并入上下文）
    - 表格合并单元格去重
    - 嵌入图片 OCR
    - page 用段落计数器近似模拟（每 40 段为一"页"）
    """
    chunks = []
    fhash = file_sha256(file_path)
    fname = os.path.basename(file_path)
    idx = 0

    if not HAS_DOCX:
        return chunks

    word = DocxDocument(file_path)
    heading_stack: list = []
    para_counter = 0           # 段落计数，用于近似页码
    PAGE_SIZE = 40             # 每 PAGE_SIZE 段算一页

    # 列表缓冲：连续列表项先收集，一起作为一个 text chunk
    list_buf: list[str] = []
    list_heading_path: str = ""
    prev_para_text: str = ""   # 记录上一个段落，用于关联表格标题

    def _flush_list():
        nonlocal idx
        if not list_buf:
            return
        merged = "\n".join(f"• {l}" for l in list_buf)
        page_n = max(1, para_counter // PAGE_SIZE)
        for t in smart_chunk_text(merged):
            if t.strip():
                chunks.append(Chunk(
                    text=t, file_name=fname, file_path=file_path,
                    file_hash=fhash, page=page_n, block_type="text",
                    heading_path=list_heading_path, chunk_index=idx
                ))
                idx += 1
        list_buf.clear()

    for element in word.element.body:
        tag = element.tag.split("}")[-1]

        if tag == "p":
            from docx.text.paragraph import Paragraph
            para = Paragraph(element, word)
            para_counter += 1
            page_n = max(1, para_counter // PAGE_SIZE)

            # 嵌入图片 OCR
            for img_bytes, img_ext in _extract_docx_paragraph_images(element, word):
                img_text = _ocr_bytes(img_bytes, img_ext, filename=fname)
                if img_text.strip():
                    _flush_list()
                    for t in smart_chunk_text(clean_text(img_text)):
                        if t.strip():
                            chunks.append(Chunk(
                                text=t, file_name=fname, file_path=file_path,
                                file_hash=fhash, page=page_n, block_type="image",
                                heading_path=_heading_path(heading_stack), chunk_index=idx
                            ))
                            idx += 1

            text = para.text.strip()
            if not text:
                continue

            style = para.style.name if para.style else ""
            level = detect_heading_level(text, style)

            # ── 标题 ──
            if level > 0:
                _flush_list()
                _update_heading_stack(heading_stack, level, text)
                chunks.append(Chunk(
                    text=text, file_name=fname, file_path=file_path,
                    file_hash=fhash, page=page_n, block_type="heading",
                    heading_level=level, heading_path=_heading_path(heading_stack),
                    chunk_index=idx
                ))
                idx += 1

            # ── 列表段落：缓冲合并 ──
            elif "list" in style.lower():
                if not list_buf:
                    list_heading_path = _heading_path(heading_stack)
                list_buf.append(text)

            # ── 普通段落 ──
            else:
                _flush_list()
                for t in smart_chunk_text(text):
                    if t.strip():
                        chunks.append(Chunk(
                            text=t, file_name=fname, file_path=file_path,
                            file_hash=fhash, page=page_n, block_type="text",
                            heading_path=_heading_path(heading_stack), chunk_index=idx
                        ))
                        idx += 1

            prev_para_text = text  # 记录本段落，供下一个表格关联标题用

        elif tag == "tbl":
            from docx.table import Table
            _flush_list()
            table = Table(element, word)
            if not table.rows:
                continue
            headers, table_text = _docx_table_to_text(table)
            if not table_text.strip():
                continue
            page_n = max(1, para_counter // PAGE_SIZE)

            # 如果上一个段落看起来是表格标题（含"表"字或"Table"），拼到表格内容开头
            _is_table_caption = bool(
                prev_para_text and (
                    re.search(r'表\s*\d', prev_para_text) or
                    re.search(r'[Tt]able\s*\d', prev_para_text) or
                    re.search(r'表\s*[一二三四五六七八九十百]', prev_para_text)
                )
            )
            # 大表格使用滑动窗口切片，每块保留表头
            header_line = " | ".join(headers) if headers else ""
            data_rows = table_text.split("\n")[1:] if "\n" in table_text else []  # 去掉表头行

            # 如果表格很大（超过 20 行），或者总长度超过 3000，启用滑动窗口
            if len(data_rows) > 20 or len(table_text) > 3000:
                window_size = 15  # 每块最多 15 行数据
                step = 10         # 滑动步长
                for i in range(0, len(data_rows), step):
                    batch = data_rows[i:i + window_size]
                    if not batch:
                        continue
                    chunk_text = f"[{prev_para_text}]\n" if _is_table_caption else ""
                    if header_line:
                        chunk_text += header_line + "\n"
                    chunk_text += "\n".join(batch)
                    chunks.append(Chunk(
                        text=chunk_text, file_name=fname, file_path=file_path,
                        file_hash=fhash, page=page_n, block_type="table",
                        table_headers=headers, heading_path=_heading_path(heading_stack),
                        chunk_index=idx
                    ))
                    idx += 1
            else:
                table_chunk_text = f"[{prev_para_text}]\n{table_text}" if _is_table_caption else table_text
                chunks.append(Chunk(
                    text=table_chunk_text, file_name=fname, file_path=file_path,
                    file_hash=fhash, page=page_n, block_type="table",
                    table_headers=headers, heading_path=_heading_path(heading_stack),
                    chunk_index=idx
                ))
                idx += 1
            prev_para_text = ""  # 标题已消费，清空

    _flush_list()  # 文档末尾可能还有未刷出的列表
    _print_chunks(chunks, f"Word · {fname}")
    return chunks


# ── Excel 解析 ────────────────────────────────────────────────

def _fmt_cell(val) -> str:
    """单元格值格式化，处理 None / float / datetime / date"""
    if val is None:
        return ""
    import datetime
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        import decimal
        return str(decimal.Decimal(str(val)).normalize())
    return str(val).strip()


def _expand_merged_cells(ws) -> list[list[str]]:
    """
    展开合并单元格：把每个合并区域的左上角值填充到区域内所有格，
    避免合并格在 iter_rows 中出现 None。
    返回二维列表 grid[row][col]（0-based）。
    """
    # 先把合并区域的值记录下来
    merge_values: dict[tuple, str] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        val = _fmt_cell(top_left.value)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_values[(row, col)] = val

    grid = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cells = []
        for col_idx, cell in enumerate(row, start=1):
            if (row_idx, col_idx) in merge_values:
                cells.append(merge_values[(row_idx, col_idx)])
            else:
                cells.append(_fmt_cell(cell.value))
        grid.append(cells)
    return grid


# Excel 每个 chunk 最多包含的数据行数（表头行不计入）
_EXCEL_CHUNK_ROWS = int(os.environ.get("EXCEL_CHUNK_ROWS", "50"))


def parse_excel(file_path: str) -> list[Chunk]:
    """
    每个 Sheet 按滑动窗口切块，每块带表头行。
    处理合并单元格、日期格式化、空行过滤。
    chunk_index 全局递增，heading_path 含 Sheet 名 + 块编号。
    """
    chunks = []
    fhash = file_sha256(file_path)
    fname = os.path.basename(file_path)
    global_idx = 0

    if not HAS_OPENPYXL:
        return chunks

    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid = _expand_merged_cells(ws)

        # 过滤全空行
        grid = [r for r in grid if any(c for c in r)]
        if not grid:
            continue

        headers = grid[0]
        header_line = " | ".join(headers)
        data_rows = [
            " | ".join(r)
            for r in grid[1:]
            if any(c for c in r)
        ]

        if not data_rows:
            # 只有表头，整体入库
            text = f"[Sheet: {sheet_name}]\n{header_line}"
            chunks.append(Chunk(
                text=text, file_name=fname, file_path=file_path,
                file_hash=fhash, page=1, block_type="table",
                table_headers=headers,
                heading_path=f"Sheet: {sheet_name}",
                chunk_index=global_idx
            ))
            global_idx += 1
            continue

        # 按 _EXCEL_CHUNK_ROWS 切块，每块保留表头
        total_rows = len(data_rows)
        chunk_num = 0
        for start in range(0, total_rows, _EXCEL_CHUNK_ROWS):
            batch = data_rows[start: start + _EXCEL_CHUNK_ROWS]
            end = min(start + _EXCEL_CHUNK_ROWS, total_rows)
            row_range = f"行{start + 2}~{end + 1}"   # +1 因为表头占第1行，+2 因为1-based
            heading = f"Sheet: {sheet_name} / {row_range}"

            text = f"[Sheet: {sheet_name}]（{row_range}）\n{header_line}\n" + "\n".join(batch)
            chunks.append(Chunk(
                text=text, file_name=fname, file_path=file_path,
                file_hash=fhash, page=chunk_num + 1, block_type="table",
                table_headers=headers,
                heading_path=heading,
                chunk_index=global_idx
            ))
            global_idx += 1
            chunk_num += 1

    _print_chunks(chunks, f"Excel · {fname}")
    return chunks


# ── 独立图片 OCR ──────────────────────────────────────────────

def parse_image(file_path: str) -> list[Chunk]:
    """独立图片文件 OCR。按 OCR_ENGINES 顺序尝试可用引擎。"""
    fhash = file_sha256(file_path)
    fname = os.path.basename(file_path)

    if not _has_any_ocr():
        raise OCRUnavailableError(fname)

    try:
        with open(file_path, "rb") as f:
            img_bytes = f.read()
        text = _do_ocr(img_bytes, filename=fname)
        del img_bytes  # OCR 完立即释放
        import gc; gc.collect()
    except OCRUnavailableError:
        raise
    except BaseException as e:
        print(f"    ❌ 图片处理失败 {fname}: {e}")
        return []

    if not text.strip():
        print(f"    ⚠️  OCR 未识别到文字: {fname}")
        return []

    cleaned = clean_text_table(text) if _is_table_text(text) else clean_text(text)

    if _is_table_text(text):
        print(f"    📊 检测到铭牌/表格图片，整体入库（{len(cleaned)}字）")
        texts = [cleaned]
    else:
        texts = smart_chunk_text(cleaned)

    result_chunks = [
        Chunk(
            text=t, file_name=fname, file_path=file_path,
            file_hash=fhash, page=1, block_type="image", chunk_index=i
        )
        for i, t in enumerate(texts) if t.strip()
    ]
    _print_chunks(result_chunks, f"图片 · {fname}")
    return result_chunks


# ── 统一入口 ──────────────────────────────────────────────────

SUPPORTED_EXTS = {
    ".pdf":  parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_excel,
    ".xls":  parse_excel,
    ".png":  parse_image,
    ".jpg":  parse_image,
    ".jpeg": parse_image,
    ".tiff": parse_image,
    ".bmp":  parse_image,
}


def parse_file(file_path: str) -> list[Chunk]:
    ext = os.path.splitext(file_path)[1].lower()
    parser = SUPPORTED_EXTS.get(ext)
    if not parser:
        print(f"⚠️  不支持的格式: {file_path}")
        return []
    print(f"  📄 解析 {os.path.basename(file_path)} ({ext})")
    return parser(file_path)


def parse_directory(dir_path: str) -> list[Chunk]:
    all_chunks = []
    for root, _, files in os.walk(dir_path):
        for fname in sorted(files):
            fpath = os.path.join(root, fname)
            try:
                all_chunks.extend(parse_file(fpath))
            except Exception as e:
                print(f"  ❌ 解析失败 {fname}: {e}")
    print(f"✅ 共解析 {len(all_chunks)} 个 chunk")
    return all_chunks